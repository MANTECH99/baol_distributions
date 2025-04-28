import logging

from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl.workbook import Workbook
from django.http import HttpResponse
from .models import Livraison, Camion, StatutCamion
from .forms import LivraisonForm
from django.db.models import Sum, Min
from django.utils.timezone import now
import json
from datetime import datetime, date, timedelta
from django.views.decorators.clickjacking import xframe_options_exempt
@xframe_options_exempt
def ajouter_livraison(request):
    # Récupérer les paramètres depuis l'URL
    date_param = request.GET.get('date')
    camion_id = request.GET.get('camion')

    try:
        selected_date = datetime.strptime(date_param, "%Y-%m-%d").date() if date_param else None
    except ValueError:
        selected_date = None

    camion = get_object_or_404(Camion, id=camion_id) if camion_id else None

    if request.method == "POST":
        form = LivraisonForm(request.POST)
        if form.is_valid():
            livraison = form.save(commit=False)
            livraison.date = selected_date
            livraison.camion = camion
            livraison.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Livraison enregistrée avec succès',
                    'livraison': {
                        'id': livraison.id,
                        'date': livraison.date.strftime('%Y-%m-%d'),
                        'date_display': livraison.date.strftime('%d %B %Y'),
                        'camion_id': camion.id,
                        'camion_numero': camion.numero,
                        'tonnage': str(livraison.tonnage),
                        'prix_unitaire': str(livraison.prix_unitaire),
                        'quantite': livraison.quantite,
                        'montant': str(livraison.montant),
                        'chiffonage': livraison.chiffonage,
                        'numero_bl': livraison.numero_bl,
                        'statut': livraison.get_statut_display(),
                    }
                })
            else:
                return HttpResponse(f"""
                    <script>
                        window.top.location.reload(true);
                    </script>
                """)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {f: [str(e) for e in form.errors[f]] for f in form.errors}
                return JsonResponse({
                    'success': False,
                    'errors': errors
                }, status=400)
            else:
                return render(request, 'gestion/ajouter_livraison.html', {
                    'form': form,
                    'camion': camion,
                    'selected_date': selected_date
                })

    # GET request - afficher le formulaire
    form = LivraisonForm()
    return render(request, 'gestion/ajouter_livraison.html', {
        'form': form,
        'camion': camion,
        'selected_date': selected_date
    })

from django.shortcuts import render
from datetime import date, datetime, timedelta
from collections import defaultdict
from .models import Livraison, Camion, StatutCamion

def liste_livraisons(request):
    # Filtres GET
    filter_type = request.GET.get('filter_type')
    selected_date = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    selected_month = request.GET.get('month', date.today().strftime('%Y-%m'))
    selected_camion_id = request.GET.get('camion')
    tonnage_filter = request.GET.get('tonnage_filter')

    selected_date_obj = None
    date_range = []

    # Base queryset optimisé
    livraisons = Livraison.objects.select_related('camion')

    # 🔁 Filtres temporels
    if filter_type == 'date':
        try:
            selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            selected_date_obj = date.today()
        date_range = [selected_date_obj]
        livraisons = livraisons.filter(date=selected_date_obj)
        selected_month = None

    elif filter_type == 'month':
        try:
            year, month = map(int, selected_month.split('-'))
            selected_date_obj = date(year, month, 1)
            next_month = selected_date_obj.replace(day=28) + timedelta(days=4)
            last_day = (next_month - timedelta(days=next_month.day)).day
            date_range = [date(year, month, day) for day in range(1, last_day + 1)]
            livraisons = livraisons.filter(date__year=year, date__month=month)
        except ValueError:
            selected_date_obj = date.today()
            livraisons = livraisons.filter(date=selected_date_obj)
            date_range = [selected_date_obj]
        selected_date = None

    else:
        selected_date_obj = date.today()
        selected_date = selected_date_obj.strftime('%Y-%m-%d')
        date_range = [selected_date_obj]
        livraisons = livraisons.filter(date=selected_date_obj)
        selected_month = None

    # Filtres supplémentaires
    if selected_camion_id and selected_camion_id != "None":
        livraisons = livraisons.filter(camion_id=selected_camion_id)

    if tonnage_filter == 'lt5':
        livraisons = livraisons.filter(tonnage__lt=5)

    livraisons = livraisons.select_related('camion').order_by('-date')

    # ✅ Préchargement mémoire
    livraisons_dict = defaultdict(lambda: defaultdict(list))
    for livraison in livraisons:
        livraisons_dict[livraison.camion_id][livraison.date].append(livraison)

    # 🔄 Récupération des camions
    camions = Camion.objects.only('id', 'numero')

    # ✅ Précharger les statuts utiles
    statuts_qs = StatutCamion.objects.filter(date__in=date_range).select_related('camion')
    statuts_dict = defaultdict(dict)
    for s in statuts_qs:
        statuts_dict[s.camion_id][s.date] = s.get_statut_display()

    # Construction des livraisons et statuts par camion et jour
    livraisons_par_camion_et_date = {}
    statuts_par_date = {}

    for camion in camions:
        livraisons_par_camion_et_date[camion.id] = {}
        statuts_par_date[camion.id] = {}

        for day in date_range:
            livraisons_jour = livraisons_dict.get(camion.id, {}).get(day)
            livraisons_par_camion_et_date[camion.id][day] = livraisons_jour if livraisons_jour else None

            statut = statuts_dict.get(camion.id, {}).get(day, "En attente")
            statuts_par_date[camion.id][day] = statut

    # Contexte
    context = {
        'statuts_par_date': statuts_par_date,
        'livraisons': livraisons,
        'camions': camions,
        'livraisons_par_camion_et_date': livraisons_par_camion_et_date,
        'selected_date': selected_date,
        'selected_month': selected_month,
        'selected_camion_id': selected_camion_id,
        'selected_date_obj': selected_date_obj,
        'filter_type': filter_type,
        'date_range': date_range,
        'tonnage_filter': tonnage_filter,
    }

    return render(request, 'gestion/livraisons.html', context)


def modifier_statut_camion(request, camion_id):
    camion = get_object_or_404(Camion, id=camion_id)
    date = request.GET.get('date')

    if request.method == "POST":
        statut = request.POST.get("statut")
        statut_camion, created = StatutCamion.objects.get_or_create(
            camion=camion,
            date=date,
            defaults={'statut': statut}
        )
        if not created:
            statut_camion.statut = statut
            statut_camion.save()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        return redirect("liste_livraisons")

    # Passez la date et les choix de statut au template
    STATUT_CHOICES = [
        ("En attente", "En attente"),
        ("En panne", "En panne"),
        ("Travaille pas", "Travaille pas"),
    ]

    context = {
        "camion": camion,
        "date": date,
        "STATUT_CHOICES": STATUT_CHOICES
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, "gestion/partials/statut_camion_form.html", context)
    return render(request, "gestion/modifier_statut_camion.html", context)

from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string

def modifier_livraison(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)

    if request.method == 'POST':
        form = LivraisonForm(request.POST, instance=livraison)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})  # ✅ Succès, pas besoin de recharger le formulaire

        # Si le formulaire contient des erreurs, on renvoie le HTML mis à jour
        html = render_to_string('gestion/partials/form_modif_livraison.html', {'form': form, 'livraison': livraison},
                                request=request)
        return JsonResponse({'success': False, 'html': html})  # ❌ Erreur, recharge le formulaire

    # Si GET, on affiche le formulaire dans le modal
    form = LivraisonForm(instance=livraison)
    html = render_to_string('gestion/partials/form_modif_livraison.html', {'form': form, 'livraison': livraison},
                            request=request)
    return JsonResponse({'html': html})


from django.shortcuts import get_object_or_404

def supprimer_livraison(request, livraison_id):
    livraison = get_object_or_404(Livraison, id=livraison_id)
    livraison.delete()
    return redirect('liste_livraisons')


def dashboard(request):
    # Récupérer tous les paramètres de filtre
    filter_type = request.GET.get('filter_type')
    selected_date = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    selected_month = request.GET.get('month', date.today().strftime('%Y-%m'))
    selected_year = request.GET.get('year', str(date.today().year))
    selected_camion_id = request.GET.get('camion')

    # Initialisation
    base_query = Livraison.objects.all()
    selected_date_obj = date.today()
    camions = Camion.objects.all()

    # Appliquer le filtre par camion en premier si sélectionné
    if selected_camion_id and selected_camion_id != "None":
        base_query = base_query.filter(camion_id=selected_camion_id)

    # Filtrage selon le type sélectionné
    if filter_type == 'date':
        try:
            selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
            livraisons = base_query.filter(date=selected_date_obj)
        except ValueError:
            selected_date_obj = date.today()
            livraisons = base_query.filter(date=selected_date_obj)
        selected_month = None
        selected_year = None

    elif filter_type == 'month':
        try:
            year, month = map(int, selected_month.split('-'))
            selected_date_obj = date(year, month, 1)
            livraisons = base_query.filter(date__year=year, date__month=month)
        except ValueError:
            selected_date_obj = date.today()
            livraisons = base_query.filter(date=selected_date_obj)
        selected_date = None
        selected_year = None

    elif filter_type == 'year':
        try:
            year = int(selected_year)
            livraisons = base_query.filter(date__year=year)
        except ValueError:
            year = date.today().year
            livraisons = base_query.filter(date__year=year)
        selected_date = None
        selected_month = None

    else:
        # Par défaut: livraisons du jour
        livraisons = base_query.filter(date=selected_date_obj)
        selected_month = None
        selected_year = None

    # Calculer les statistiques globales
    stats = livraisons.aggregate(
        total_tonnage=Sum('tonnage'),
        total_montant=Sum('montant')
    )
    total_tonnage = stats['total_tonnage'] or 0
    total_montant = stats['total_montant'] or 0
    total_livraisons = livraisons.count()

    # Préparation des données pour les graphiques
    mois_labels, tonnage_values, montant_values = [], [], []

    if filter_type == 'year':
        year = int(selected_year) if selected_year and filter_type == 'year' else date.today().year
        for mois in range(1, 13):
            query = base_query.filter(date__year=year, date__month=mois)
            stats_mois = query.aggregate(
                tonnage_mois=Sum('tonnage'),
                montant_mois=Sum('montant')
            )
            mois_labels.append(f"{mois}/{year}")
            tonnage_values.append(float(stats_mois['tonnage_mois'] or 0))
            montant_values.append(float(stats_mois['montant_mois'] or 0))

    elif filter_type == 'month':
        try:
            year, month = map(int, selected_month.split('-'))
            from calendar import monthrange
            nb_jours = monthrange(year, month)[1]
            for jour in range(1, nb_jours + 1):
                query = base_query.filter(date__year=year, date__month=month, date__day=jour)
                stats_jour = query.aggregate(
                    tonnage_jour=Sum('tonnage'),
                    montant_jour=Sum('montant')
                )
                mois_labels.append(f"{jour}/{month}")
                tonnage_values.append(float(stats_jour['tonnage_jour'] or 0))
                montant_values.append(float(stats_jour['montant_jour'] or 0))
        except ValueError:
            pass

    else:
        # Par défaut: 6 derniers mois
        mois_actuel = date.today().month
        annee_actuelle = date.today().year
        for i in range(6):
            mois = (mois_actuel - i) % 12 or 12
            annee = annee_actuelle if mois_actuel - i > 0 else annee_actuelle - 1
            query = base_query.filter(date__year=annee, date__month=mois)
            stats_mois = query.aggregate(
                tonnage_mois=Sum('tonnage'),
                montant_mois=Sum('montant')
            )
            mois_labels.append(f"{mois}/{annee}")
            tonnage_values.append(float(stats_mois['tonnage_mois'] or 0))
            montant_values.append(float(stats_mois['montant_mois'] or 0))

    # Récupérer les livraisons filtrées
    livraisons_mois = livraisons.order_by('-date')[:50].values(
        'date', 'camion__numero', 'camion_id', 'camion__telephone', 'tonnage',
        'prix_unitaire', 'quantite', 'montant', 'chiffonage', 'numero_bl', 'statut'
    )

    # Contexte
    context = {
        'selected_date': selected_date,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_camion_id': selected_camion_id,
        'selected_date_obj': selected_date_obj,
        'filter_type': filter_type,
        'total_tonnage': total_tonnage,
        'total_montant': total_montant,
        'total_livraisons': total_livraisons,
        'mois_labels': json.dumps(mois_labels[::-1]),
        'tonnage_values': json.dumps(tonnage_values[::-1]),
        'montant_values': json.dumps(montant_values[::-1]),
        'livraisons_mois': livraisons_mois,
        'last_five_years': [date.today().year - i for i in range(5)],
        'camions': camions,
    }

    return render(request, 'gestion/dashboard.html', context)

from django.utils import timezone
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, date
from django.http import HttpResponse
from collections import defaultdict
from .models import Livraison, Camion, StatutCamion
import logging

logger = logging.getLogger(__name__)

def exporter_livraisons_excel(request):
    filter_type = request.GET.get('filter_type')
    selected_date = request.GET.get('date')
    selected_month = request.GET.get('month')
    selected_camion_id = request.GET.get('camion')
    tonnage_filter = request.GET.get('tonnage_filter')

    logger.info(f"Filter type: {filter_type}, Date: {selected_date}, Month: {selected_month}, Camion: {selected_camion_id}, Tonnage: {tonnage_filter}")

    wb = Workbook()
    ws = wb.active
    headers = [
        "Date", "Camion", "Numéro Téléphone", "Tonnage", "Prix Unitaire",
        "Quantité", "Montant", "Chiffonage", "N° BL", "Statut"
    ]
    ws.append(headers)

    total_tonnage = 0
    total_montant = 0

    livraisons_dict = defaultdict(lambda: defaultdict(list))
    statuts_dict = defaultdict(dict)

    if filter_type == 'month' and selected_month:
        try:
            year, month = map(int, selected_month.split('-'))
            num_days = monthrange(year, month)[1]
            date_range = [date(year, month, day) for day in range(1, num_days + 1)]

            camions = Camion.objects.only('id', 'numero', 'telephone')
            if selected_camion_id:
                camions = camions.filter(id=selected_camion_id)
            camions = list(camions.order_by('numero'))

            livraison_qs = Livraison.objects.select_related('camion').filter(
                date__year=year, date__month=month
            )
            if selected_camion_id:
                livraison_qs = livraison_qs.filter(camion_id=selected_camion_id)
            if tonnage_filter == 'lt5':
                livraison_qs = livraison_qs.filter(tonnage__lt=5)

            for l in livraison_qs.order_by('camion__numero', 'date'):
                livraisons_dict[l.camion_id][l.date].append(l)

            statut_qs = StatutCamion.objects.filter(
                date__in=date_range
            ).select_related('camion')
            for s in statut_qs:
                statuts_dict[s.camion_id][s.date] = s.get_statut_display()

            for camion in camions:
                camion_id = camion.id
                for current_date in date_range:
                    livraisons = livraisons_dict[camion_id].get(current_date)
                    if livraisons:
                        for l in livraisons:
                            ws.append([
                                l.date.strftime("%d/%m/%Y"),
                                l.camion.numero,
                                l.camion.telephone,
                                l.tonnage,
                                l.prix_unitaire,
                                l.quantite,
                                l.montant,
                                l.chiffonage,
                                l.numero_bl,
                                l.get_statut_display(),
                            ])
                            total_tonnage += l.tonnage
                            total_montant += l.montant
                    elif tonnage_filter != 'lt5':
                        statut = statuts_dict[camion_id].get(current_date, "En attente")
                        row_num = ws.max_row + 1
                        # Ajouter la ligne avec "Aucune livraison" fusionnée
                        ws.append([
                            current_date.strftime("%d/%m/%Y"),
                            camion.numero,
                            camion.telephone,
                            "", "", "", "", "", "", statut
                        ])
                        # Fusionner les colonnes D -> I (4 à 9)
                        ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=9)
                        cell = ws.cell(row=row_num, column=4)
                        cell.value = "Aucune livraison"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        statut_cell = ws.cell(row=ws.max_row, column=10)
                        fill_colors = {
                            "En panne": "FF9999",
                            "Travaille pas": "FFF699",
                            "En attente": "B4C6E7"
                        }
                        color = fill_colors.get(statut)
                        if color:
                            statut_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        except ValueError:
            logger.error("Format de mois invalide")
            return HttpResponse("Format de mois invalide")

    else:
        livraisons = Livraison.objects.select_related('camion')

        if selected_camion_id:
            livraisons = livraisons.filter(camion_id=selected_camion_id)
        if filter_type == 'date' and selected_date:
            try:
                date_obj = datetime.strptime(selected_date, "%Y-%m-%d").date()
                livraisons = livraisons.filter(date=date_obj)
            except ValueError:
                logger.error("Format de date invalide")
                livraisons = Livraison.objects.none()
        if tonnage_filter == 'lt5':
            livraisons = livraisons.filter(tonnage__lt=5)

        for l in livraisons.order_by('camion__numero', 'date'):
            ws.append([
                l.date.strftime("%d/%m/%Y"),
                l.camion.numero,
                l.camion.telephone,
                l.tonnage,
                l.prix_unitaire,
                l.quantite,
                l.montant,
                l.chiffonage,
                l.numero_bl,
                l.get_statut_display(),
            ])
            total_tonnage += l.tonnage
            total_montant += l.montant

    ws.append([])
    ws.append(["Total", "", "", total_tonnage, "", "", total_montant, "", "", ""])

    filename_parts = []
    if selected_camion_id:
        filename_parts.append(f"camion_{selected_camion_id}")
    if selected_date:
        filename_parts.append(f"date_{selected_date}")
    elif selected_month:
        filename_parts.append(f"mois_{selected_month}")
    if tonnage_filter == 'lt5':
        filename_parts.append("tonnage_lt5")

    filename = f"livraisons_{'_'.join(filename_parts) or 'all'}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response




# Vue pour l'inscription
def register(request):
    error = None
    if request.method == "POST":
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if password1 != password2:
            error = "Les mots de passe ne correspondent pas."
        elif User.objects.filter(username=username).exists():
            error = "Ce nom d'utilisateur existe déjà."
        else:
            user = User.objects.create_user(username=username, password=password1)
            login(request, user)
            return redirect("login")  # Redirige vers le tableau de bord

    return render(request, "registration/login_register.html", {"register_error": error})

# Vue pour la connexion
def user_login(request):
    error = None
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")  # Redirige vers le tableau de bord
        else:
            error = "Identifiant ou mot de passe incorrect."

    return render(request, "registration/login_register.html", {"login_error": error})




from django.shortcuts import render, redirect
from .models import Rapport
from django.http import FileResponse
from .forms import RapportForm


def liste_rapports(request):
    mois_selectionne = request.GET.get('mois')
    if mois_selectionne:
        rapports = Rapport.objects.filter(mois__month=mois_selectionne.split('-')[1],
                                          mois__year=mois_selectionne.split('-')[0])
    else:
        rapports = Rapport.objects.all()

    return render(request, 'gestion/rapports.html', {'rapports': rapports})


def telecharger_rapport(request, rapport_id):
    rapport = Rapport.objects.get(id=rapport_id)
    return FileResponse(rapport.fichier.open('rb'), as_attachment=True, filename=rapport.fichier.name)


from datetime import datetime
from django.http import JsonResponse


def ajouter_rapport(request):
    if request.method == 'POST':
        print("📥 Requête POST reçue")
        print("🔍 Données POST :", request.POST)

        # Vérifier si 'mois' est présent et le reformater
        mois_str = request.POST.get('mois')
        if mois_str:
            try:
                # Transformer 'YYYY-MM' en 'YYYY-MM-01' pour être une date valide
                mois_corrige = datetime.strptime(mois_str + '-01', '%Y-%m-%d').date()

                # Remplacer la valeur POST (hack Django pour éviter l'erreur)
                post_data = request.POST.copy()
                post_data['mois'] = mois_corrige

                # Initialiser le formulaire avec la valeur corrigée
                form = RapportForm(post_data, request.FILES)

                if form.is_valid():
                    form.save()
                    print("✅ Rapport enregistré avec succès")
                    return redirect('liste_rapports')
                else:
                    print("❌ Formulaire invalide :", form.errors)
                    return JsonResponse({"error": form.errors}, status=400)

            except ValueError as e:
                print(f"❌ Erreur conversion date : {e}")
                return JsonResponse({"error": f"Erreur conversion date : {str(e)}"}, status=400)

        else:
            print("⚠ ERREUR : 'mois' est vide")
            return JsonResponse({"error": "Le champ 'mois' est vide"}, status=400)

    else:
        form = RapportForm()

    return render(request, 'gestion/ajouter_rapport.html', {'form': form})


from django.shortcuts import get_object_or_404

def supprimer_rapport(request, rapport_id):
    rapport = get_object_or_404(Rapport, id=rapport_id)
    rapport.delete()
    return redirect('liste_rapports')
